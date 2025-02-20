"""
用于处理excel,将红色字符替换为"\"
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def replace_red_text(input_file, output_file):
    # Load the workbook and select the first sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Iterate over all cells in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell has a font and if its color is red
            if cell.font and cell.font.color and cell.font.color.rgb == "FFFF0000":
                # Replace the cell value with a red backslash
                cell.value = "\\"
                cell.font = Font(color="FFFF0000")

    # Save the workbook to a new file
    wb.save(output_file)

# Example usage
input_file = "RQ1表格.xlsx"
output_file = "output.xlsx"
replace_red_text(input_file, output_file)
print(f"处理完成，结果已保存到 {output_file}")
